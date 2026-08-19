"""
יצירת דוח אקסל מלא - גיליון נפרד לכל מחלקה (6 גיליונות)
כל שורה = חייל, כל עמודה = סוג ציוד, התא = הכמות שהוא מחזיק (ריק אם אין)
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from db.operations import get_full_report_matrix

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def build_full_report(output_path: str, layer: str = "personal"):
    item_names, platoons_data = get_full_report_matrix(layer)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # מסירים את הגיליון הריק שנוצר כברירת מחדל

    for platoon in platoons_data:
        # שם גיליון מוגבל ל-31 תווים באקסל - השמות שלנו קצרים, לא בעיה בפועל
        ws = wb.create_sheet(title=platoon["platoon_name"][:31])
        ws.sheet_view.rightToLeft = True

        headers = ["שם חייל"] + item_names
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for soldier in platoon["soldiers"]:
            row = [soldier["full_name"]] + [soldier["items"].get(name, "") for name in item_names]
            ws.append(row)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")
                if cell.column > 1:
                    cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 22
        for idx in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 13

        ws.freeze_panes = "B2"

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "full_report.xlsx"
    build_full_report(out)
    print(f"נשמר: {out}")
